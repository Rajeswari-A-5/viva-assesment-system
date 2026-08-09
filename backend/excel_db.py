from openpyxl import Workbook, load_workbook
from pathlib import Path
from datetime import datetime


# =========================================================
# FILE PATHS
# =========================================================

BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
EXCEL_FILE = DATA_DIR / "VivaTrack.xlsx"


# =========================================================
# CREATE EXCEL FILE
# =========================================================

def create_excel_file():

    DATA_DIR.mkdir(exist_ok=True)

    # If file doesn't exist, create it
    if not EXCEL_FILE.exists():

        workbook = Workbook()

        # -------------------------------------------------
        # Students
        # -------------------------------------------------

        students = workbook.active
        students.title = "Students"

        students.append([
            "Register No",
            "Name",
            "Batch",
            "Date",
            "Start Time",
            "End Time",
            "Status",
            "Score"
        ])

        # -------------------------------------------------
        # Questions
        # -------------------------------------------------

        questions = workbook.create_sheet("Questions")

        questions.append([
            "Question ID",
            "Question",
            "Type",
            "Category",
            "Difficulty",
            "Option A",
            "Option B",
            "Option C",
            "Option D",
            "Correct Answer"
        ])

        # -------------------------------------------------
        # Responses
        # -------------------------------------------------

        responses = workbook.create_sheet("Responses")

        responses.append([
            "Response ID",
            "Register No",
            "Question ID",
            "Question",
            "Type",
            "Response",
            "Correct",
            "Time Taken"
        ])

        # -------------------------------------------------
        # Admin
        # -------------------------------------------------

        admin = workbook.create_sheet("Admin")

        admin.append([
            "Email",
            "Password",
            "Name"
        ])

        admin.append([
            "admin@institution.edu",
            "admin123",
            "Lab Administrator"
        ])

        workbook.save(EXCEL_FILE)
        workbook.close()

        print("VivaTrack.xlsx created successfully.")

    else:

        # -------------------------------------------------
        # FILE ALREADY EXISTS
        # Make sure Admin sheet exists
        # -------------------------------------------------

        workbook = load_workbook(EXCEL_FILE)

        if "Admin" not in workbook.sheetnames:

            admin = workbook.create_sheet("Admin")

            admin.append([
                "Email",
                "Password",
                "Name"
            ])

            admin.append([
                "admin@institution.edu",
                "admin123",
                "Lab Administrator"
            ])

            workbook.save(EXCEL_FILE)

            print("Admin sheet added successfully.")

        else:

            admin = workbook["Admin"]

            # Check whether admin credentials exist
            admin_exists = False

            for row in admin.iter_rows(min_row=2, values_only=True):

                if (
                    row[0]
                    and str(row[0]).lower()
                    == "admin@institution.edu"
                ):

                    admin_exists = True
                    break

            if not admin_exists:

                admin.append([
                    "admin@institution.edu",
                    "admin123",
                    "Lab Administrator"
                ])

                workbook.save(EXCEL_FILE)

                print("Default admin added successfully.")

        workbook.close()


# =========================================================
# GET WORKBOOK
# =========================================================

def get_workbook():

    create_excel_file()

    return load_workbook(EXCEL_FILE)


# =========================================================
# GET ALL STUDENTS
# =========================================================

def get_students():

    workbook = get_workbook()
    sheet = workbook["Students"]

    students = []

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        if not row[0]:
            continue

        students.append({

            "registerNo": row[0],
            "name": row[1],
            "batch": row[2],
            "date": row[3],
            "startTime": row[4],
            "endTime": row[5],
            "status": row[6],
            "score": row[7]

        })

    workbook.close()

    return students


# =========================================================
# GET ALL QUESTIONS
# =========================================================

def get_questions():

    workbook = get_workbook()
    sheet = workbook["Questions"]

    questions = []

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        if not row[0]:
            continue

        questions.append({

            "id": row[0],
            "question": row[1],
            "type": row[2],
            "category": row[3],
            "difficulty": row[4],
            "optionA": row[5],
            "optionB": row[6],
            "optionC": row[7],
            "optionD": row[8],
            "correctAnswer": row[9]

        })

    workbook.close()

    return questions


# =========================================================
# ADD STUDENT
# =========================================================

def add_student(
    register_no,
    name,
    batch
):

    workbook = get_workbook()
    sheet = workbook["Students"]

    # Check duplicate register number

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        if (
            row[0]
            and str(row[0]).upper()
            == register_no.upper()
        ):

            workbook.close()

            return (
                False,
                "Student with this register number already exists."
            )

    # Current date/time

    now = datetime.now()

    date = now.strftime("%d-%m-%Y")
    start_time = now.strftime("%H:%M:%S")

    # Add student

    sheet.append([

        register_no.upper(),
        name,
        batch,
        date,
        start_time,
        "",
        "Pending",
        ""

    ])

    workbook.save(EXCEL_FILE)
    workbook.close()

    return (
        True,
        "Student registered successfully."
    )


# =========================================================
# ADD QUESTION
# =========================================================

def add_question(
    question_id,
    question,
    question_type,
    category,
    difficulty,
    option_a="",
    option_b="",
    option_c="",
    option_d="",
    correct_answer=""
):

    workbook = get_workbook()
    sheet = workbook["Questions"]

    # Check duplicate Question ID

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        if (
            row[0]
            and str(row[0]).upper()
            == question_id.upper()
        ):

            workbook.close()

            return (
                False,
                "Question ID already exists."
            )

    # Validate question type

    if question_type not in ["Text", "MCQ"]:

        workbook.close()

        return (
            False,
            "Question type must be Text or MCQ."
        )

    # Text question

    if question_type == "Text":

        option_a = ""
        option_b = ""
        option_c = ""
        option_d = ""
        correct_answer = ""

    # MCQ question

    if question_type == "MCQ":

        if not option_a or not option_b:
            workbook.close()

            return (
                False,
                "MCQ options are required."
            )

        if not option_c or not option_d:
            workbook.close()

            return (
                False,
                "MCQ options are required."
            )

        if not correct_answer:

            workbook.close()

            return (
                False,
                "Correct answer is required for MCQ."
            )

        valid_answers = [
            "A",
            "B",
            "C",
            "D"
        ]

        if correct_answer.upper() not in valid_answers:

            workbook.close()

            return (
                False,
                "Correct answer must be A, B, C or D."
            )

        correct_answer = correct_answer.upper()

    # Add question

    sheet.append([

        question_id.upper(),
        question,
        question_type,
        category,
        difficulty,
        option_a,
        option_b,
        option_c,
        option_d,
        correct_answer

    ])

    workbook.save(EXCEL_FILE)
    workbook.close()

    return (
        True,
        "Question added successfully."
    )


# =========================================================
# GET STUDENT BY REGISTER NUMBER
# =========================================================

def get_student_by_register_no(register_no):

    workbook = get_workbook()
    sheet = workbook["Students"]

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        if (
            row[0]
            and str(row[0]).upper()
            == register_no.upper()
        ):

            student = {

                "registerNo": row[0],
                "name": row[1],
                "batch": row[2],
                "date": row[3],
                "startTime": row[4],
                "endTime": row[5],
                "status": row[6],
                "score": row[7]

            }

            workbook.close()

            return student

    workbook.close()

    return None


# =========================================================
# SAVE RESPONSE
# =========================================================

def save_response(
    response_id,
    register_no,
    question_id,
    question,
    question_type,
    response,
    correct,
    time_taken
):

    workbook = get_workbook()
    sheet = workbook["Responses"]

    sheet.append([

        response_id,
        register_no,
        question_id,
        question,
        question_type,
        response,
        correct,
        time_taken

    ])

    workbook.save(EXCEL_FILE)
    workbook.close()


# =========================================================
# UPDATE STUDENT RESULT
# =========================================================

def update_student_result(
    register_no,
    end_time,
    status,
    score
):

    workbook = get_workbook()
    sheet = workbook["Students"]

    for row in sheet.iter_rows(
        min_row=2
    ):

        if (
            row[0].value
            and str(row[0].value).upper()
            == register_no.upper()
        ):

            row[5].value = end_time
            row[6].value = status
            row[7].value = score

            break

    workbook.save(EXCEL_FILE)
    workbook.close()


# =========================================================
# VERIFY ADMIN
# =========================================================

def verify_admin(
    email,
    password
):

    workbook = get_workbook()
    sheet = workbook["Admin"]

    for row in sheet.iter_rows(
        min_row=2,
        values_only=True
    ):

        stored_email = row[0]
        stored_password = row[1]
        stored_name = row[2]

        if (
            stored_email
            and stored_password
            and str(stored_email).strip().lower()
            == email.strip().lower()
            and str(stored_password).strip()
            == password.strip()
        ):

            admin = {

                "email": stored_email,
                "name": stored_name

            }

            workbook.close()

            return admin

    workbook.close()

    return None


# =========================================================
# DELETE QUESTION
# =========================================================

def delete_question(question_id):

    workbook = get_workbook()
    sheet = workbook["Questions"]

    row_to_delete = None

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

        if row[0] and str(row[0]).upper() == question_id.upper():
            row_to_delete = row_idx
            break

    if row_to_delete:
        sheet.delete_rows(row_to_delete)
        workbook.save(EXCEL_FILE)
        workbook.close()
        return True, "Question deleted successfully."

    workbook.close()
    return False, "Question not found."


# =========================================================
# DELETE STUDENT
# =========================================================

def delete_student(register_no):

    workbook = get_workbook()
    sheet = workbook["Students"]

    row_to_delete = None

    for row_idx, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):

        if row[0] and str(row[0]).upper() == register_no.upper():
            row_to_delete = row_idx
            break

    if row_to_delete:
        sheet.delete_rows(row_to_delete)
        workbook.save(EXCEL_FILE)
        workbook.close()
        return True, "Student deleted successfully."

    workbook.close()
    return False, "Student not found."


# =========================================================
# GET STUDENT RESPONSES FOR EVALUATION
# =========================================================

def get_student_responses(register_no):

    workbook = get_workbook()
    sheet = workbook["Responses"]

    responses = []

    for row in sheet.iter_rows(min_row=2, values_only=True):

        if row[1] and str(row[1]).upper() == register_no.upper():

            responses.append({
                "responseId": row[0],
                "registerNo": row[1],
                "questionId": row[2],
                "question": row[3],
                "questionType": row[4],
                "response": row[5],
                "correct": row[6],
                "timeTaken": row[7]
            })

    workbook.close()

    return responses


# =========================================================
# GRADE / EVALUATE STUDENT RESPONSES
# =========================================================

def grade_student_responses(register_no, score, remarks=None, status="Completed"):

    workbook = get_workbook()
    sheet = workbook["Students"]

    updated = False

    for row in sheet.iter_rows(min_row=2):

        if row[0].value and str(row[0].value).upper() == register_no.upper():

            now_time = datetime.now().strftime("%H:%M:%S")
            if not row[5].value:
                row[5].value = now_time

            row[6].value = status
            row[7].value = score

            updated = True
            break

    workbook.save(EXCEL_FILE)
    workbook.close()

    if updated:
        return True, "Student evaluated and score updated successfully."
    return False, "Student record not found."


# =========================================================
# GET DASHBOARD STATS
# =========================================================

def get_dashboard_stats():

    students = get_students()
    questions = get_questions()

    total_students = len(students)
    total_questions = len(questions)

    completed = sum(1 for s in students if str(s.get("status", "")).lower() == "completed")
    pending = sum(1 for s in students if str(s.get("status", "")).lower() == "pending")

    scores = []
    for s in students:
        sc = s.get("score")
        if sc is not None and sc != "":
            try:
                scores.append(float(sc))
            except ValueError:
                pass

    avg_score = round(sum(scores) / len(scores), 1) if scores else 0.0

    return {
        "totalStudents": total_students,
        "totalQuestions": total_questions,
        "completed": completed,
        "pending": pending,
        "avgScore": avg_score
    }


# =========================================================
# SEED DEFAULT VIVA QUESTIONS
# =========================================================

def seed_default_questions():

    questions = get_questions()
    if len(questions) >= 5:
        return

    sample_questions = [
        ("Q101", "What is the primary difference between a list and a tuple in Python?", "MCQ", "Python", "Easy", "Lists are immutable, tuples are mutable", "Lists are mutable, tuples are immutable", "Lists store strings only, tuples store numbers only", "There is no difference", "B"),
        ("Q102", "Which HTTP method is used to update an existing resource or create a new resource if it doesn't exist?", "MCQ", "Web Dev", "Medium", "GET", "DELETE", "PUT", "OPTIONS", "C"),
        ("Q103", "Explain the concept of Object-Oriented Programming (OOP) encapsulation with a practical example.", "Text", "OOP", "Medium", "", "", "", "", ""),
        ("Q104", "What is the time complexity of binary search on a sorted array of size n?", "MCQ", "Data Structures", "Easy", "O(n)", "O(n^2)", "O(log n)", "O(1)", "C"),
        ("Q105", "Describe how Virtual DOM works in modern web frameworks and why it improves rendering efficiency.", "Text", "Web Dev", "Hard", "", "", "", "", ""),
        ("Q106", "Which SQL clause is used to filter records after an aggregation operation?", "MCQ", "Databases", "Medium", "WHERE", "HAVING", "GROUP BY", "ORDER BY", "B"),
        ("Q107", "What is a closure in JavaScript? Explain how lexical scope enables closures.", "Text", "JavaScript", "Medium", "", "", "", "", ""),
        ("Q108", "Which data structure operates on a First-In, First-Out (FIFO) principle?", "MCQ", "Data Structures", "Easy", "Stack", "Tree", "Queue", "Graph", "C")
    ]

    for q in sample_questions:
        add_question(q[0], q[1], q[2], q[3], q[4], q[5], q[6], q[7], q[8], q[9])

    print("Default viva questions seeded successfully.")


# =========================================================
# PDF QUESTION BANK PARSER
# =========================================================

def parse_pdf_questions(pdf_stream):
    try:
        from pypdf import PdfReader
        import re

        reader = PdfReader(pdf_stream)
        full_text = ""
        for page in reader.pages:
            txt = page.extract_text()
            if txt:
                full_text += txt + "\n"

        lines = [line.strip() for line in full_text.splitlines() if line.strip()]
        parsed_questions = []
        current_q = None

        for line in lines:
            q_match = re.match(r'^(?:Q\d+[\.:]?|\d+[\.:])\s*(.+)', line, re.IGNORECASE)
            if q_match:
                if current_q:
                    parsed_questions.append(current_q)
                q_text = q_match.group(1).strip()
                current_q = {
                    "id": f"Q{len(parsed_questions) + 1:02d}",
                    "question": q_text,
                    "type": "Text",
                    "category": "Lab Viva",
                    "difficulty": "Medium",
                    "optionA": "",
                    "optionB": "",
                    "optionC": "",
                    "optionD": "",
                    "correctAnswer": ""
                }
                continue

            if current_q:
                opt_match = re.match(r'^(?:[A-D][\.\)]|\([A-D]\))\s*(.+)', line, re.IGNORECASE)
                ans_match = re.match(r'^(?:Answer|Correct Answer|Ans)[\s:]*([A-D])', line, re.IGNORECASE)

                if ans_match:
                    current_q["correctAnswer"] = ans_match.group(1).upper()
                    current_q["type"] = "MCQ"
                elif opt_match:
                    opt_letter = line[0].upper()
                    if opt_letter in ["A", "B", "C", "D"]:
                        current_q[f"option{opt_letter}"] = opt_match.group(1).strip()
                        current_q["type"] = "MCQ"
                else:
                    if not current_q["optionA"]:
                        current_q["question"] += " " + line

        if current_q:
            parsed_questions.append(current_q)

        return parsed_questions
    except Exception as e:
        print("PDF Parsing error:", e)
        return []


# =========================================================
# PER-LAB EXCEL SHEET MANAGEMENT
# =========================================================

def save_lab_session_result(register_no, name, batch, group, experiment, date_str, start_time, end_time, status, mcq_score, final_score, remarks=""):

    workbook = get_workbook()

    b_str = str(batch).replace("Batch", "").strip()
    g_str = str(group).replace("Group", "").strip()
    d_str = str(date_str).replace("-", "").replace("/", "")

    sheet_name = f"Lab_B{b_str}_G{g_str}_{d_str}"[:31]


    if sheet_name not in workbook.sheetnames:
        sheet = workbook.create_sheet(sheet_name)
        sheet.append([
            "Register No",
            "Name",
            "Batch",
            "Group",
            "Experiment",
            "Date",
            "Start Time",
            "End Time",
            "Status",
            "MCQ Score",
            "Final Score",
            "Remarks"
        ])
    else:
        sheet = workbook[sheet_name]

    existing_row = None
    for row in sheet.iter_rows(min_row=2):
        if row[0].value and str(row[0].value).upper() == register_no.upper():
            existing_row = row
            break

    if existing_row:
        existing_row[5].value = date_str
        existing_row[6].value = start_time
        existing_row[7].value = end_time
        existing_row[8].value = status
        existing_row[9].value = mcq_score
        existing_row[10].value = final_score
        existing_row[11].value = remarks
    else:
        sheet.append([
            register_no.upper(),
            name,
            batch,
            group,
            experiment,
            date_str,
            start_time,
            end_time,
            status,
            mcq_score,
            final_score,
            remarks
        ])

    workbook.save(EXCEL_FILE)
    workbook.close()
    return sheet_name


def get_all_lab_sheets():

    workbook = get_workbook()
    lab_sheets = []
    for name in workbook.sheetnames:
        if name.startswith("Lab_"):
            sheet = workbook[name]
            rows_count = max(0, sheet.max_row - 1)
            lab_sheets.append({
                "sheetName": name,
                "studentCount": rows_count
            })
    workbook.close()
    return lab_sheets


def get_lab_sheet_data(sheet_name):

    workbook = get_workbook()
    if sheet_name not in workbook.sheetnames:
        workbook.close()
        return None

    sheet = workbook[sheet_name]
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        data.append({
            "registerNo": row[0],
            "name": row[1],
            "batch": row[2],
            "group": row[3],
            "experiment": row[4],
            "date": row[5],
            "startTime": row[6],
            "endTime": row[7],
            "status": row[8],
            "mcqScore": row[9],
            "finalScore": row[10],
            "remarks": row[11]
        })
    workbook.close()
    return data



